import customtkinter as ctk
from tkinter import ttk, messagebox
from tkcalendar import Calendar
from datetime import timedelta, datetime
import os
from openpyxl import Workbook, load_workbook
from fpdf import FPDF
from PIL import Image


ctk.set_appearance_mode('dark')
ctk.set_default_color_theme('blue')


class MultiUserTimerApp:
    def __init__(self, root):
        self.root = root
        self.root.title('Orizon Timer - Multi-Utente con Data')
        self.root.geometry('1000x760')
        self.assets_folder = 'assets'
        os.makedirs(self.assets_folder, exist_ok=True)

        # Variabili di stato
        self.users = []
        self.current_date = None
        self.running_timers = {}
        self.paused_timers = {}  # Memorizza i timer in pausa

        # Interfaccia
        self.create_widgets()
        self.update_timers()

    def create_widgets(self):
        self.main_frame = ctk.CTkFrame(self.root)
        self.main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Parte sinistra
        self.left_frame = ctk.CTkFrame(self.main_frame)
        self.left_frame.pack(side='left', padx=10, pady=10, fill='y')

        ctk.CTkLabel(self.left_frame, text='Seleziona Data:').pack(pady=5)
        self.calendar = Calendar(self.left_frame, selectmode='day', date_pattern='dd/mm/yyyy')
        self.calendar.pack(pady=5)

        # Pulsanti
        ctk.CTkButton(self.left_frame, text='Conferma Data', command=self.set_date).pack(pady=5)
        ctk.CTkButton(self.left_frame, text='Esporta PDF', command=self.export_to_pdf, fg_color='#FF8C00').pack(pady=5)
    
        ctk.CTkLabel(self.left_frame, text='Nome Utente:').pack(pady=5)
        self.user_name_var = ctk.StringVar()
        ctk.CTkEntry(self.left_frame, textvariable=self.user_name_var, width=200).pack(pady=5)
        ctk.CTkButton(self.left_frame, text='Aggiungi', command=self.add_user).pack(pady=5)
        ctk.CTkButton(self.left_frame, text='Stop/Pausa', command=self.stop_timer).pack(pady=5)
        ctk.CTkButton(self.left_frame, text='Salva', command=self.save_to_excel, fg_color='#006400').pack(pady=5)
        ctk.CTkButton(self.left_frame, text='Elimina', command=self.delete_user, fg_color='#8B0000').pack(pady=5)
        
        ctk.CTkLabel(self.left_frame, text='Cerca Utente:').pack(pady=5)
        self.search_var = ctk.StringVar()
        self.search_entry = ctk.CTkEntry(self.left_frame, textvariable=self.search_var, width=200)
        self.search_entry.pack(pady=5)
        self.search_entry.bind("<KeyRelease>", self.filter_users)

                # Aggiungi il logo sotto l'input della ricerca
# Aggiungi il logo sotto l'input della ricerca
        logo_path = "logo/orizon.png"
        if os.path.exists(logo_path):
            image = Image.open(logo_path)  # Apri l'immagine con PIL
            image = image.resize((150, 150), Image.Resampling.LANCZOS)  # Ridimensiona l'immagine

    # Crea una CTkImage solo con l'immagine ridimensionata
            self.logo_image = ctk.CTkImage(light_image=image, dark_image=image)

    # Crea un widget CTkLabel con l'immagine ridimensionata
            self.logo_label = ctk.CTkLabel(self.left_frame, image=self.logo_image, text='')

    # Imposta un'ancora per il posizionamento e un padding personalizzato
            self.logo_label.pack(side='bottom', anchor='sw', pady=50, padx=100)
        else:
            print(f"Immagine non trovata: {logo_path}")




        style = ttk.Style()
        style.configure('Treeview', rowheight=30, font=('Helvetica', 16))  # Aumenta la dimensione del font
        style.configure('Treeview.Heading', font=('Helvetica', 16, 'bold'))  # Intestazioni in grassetto e più grandi
       
        # Parte destra
        self.right_frame = ctk.CTkFrame(self.main_frame)
        self.right_frame.pack(side='right', fill='both', expand=True, padx=10, pady=10)
        
                # A questo punto, puoi usare self.right_frame senza errore
        self.total_cost_var = ctk.StringVar(value="0.00")
        
        # Nuova Label per indicare "Costo Totale:"
        self.total_cost_text_label = ctk.CTkLabel(self.right_frame, text="Totale:", font=("Helvetica", 17))
        self.total_cost_text_label.pack(pady=5)

        # Label che mostra il valore totale con l'aggiornamento del valore e "EUR"
        self.total_cost_label = ctk.CTkLabel(self.right_frame, textvariable=self.total_cost_var, font=("Helvetica", 22, 'bold'))
        self.total_cost_label.pack(pady=5)


        columns = ('Nome', 'Timer', 'Costo')
        self.tree = ttk.Treeview(self.right_frame, columns=columns, show='headings', height=20)
        self.tree.heading('Nome', text='Nome')
        self.tree.heading('Timer', text='Timer')
        self.tree.heading('Costo', text='Costo')
        self.tree.column('Nome', width=200)
        self.tree.column('Timer', width=150)
        self.tree.column('Costo', width=100)
        self.tree.pack(pady=10, fill='both', expand=True)

        # Configurazione del tag 'running' dopo che il widget tree è stato creato
        self.tree.tag_configure('running', foreground='Orange')

     # Aggiungi un trace alla variabile total_cost_var
        self.total_cost_var.trace("w", lambda *args: self.update_total_cost_label())

    def update_total_cost_label(self):
        total_cost = float(self.total_cost_var.get())  # Ottieni il valore della cifra
        if total_cost > 0:
            self.total_cost_label.configure(text_color="green")  # Colore verde per il testo
        elif total_cost < 0:
            self.total_cost_label.configure(text_color="red")  # Colore rosso per il testo
        else:
            self.total_cost_label.configure(text_color="white")  # Colore bianco per il testo

        self.total_cost_var.set(f"{total_cost:.2f} EUR")



    def export_to_pdf(self):
        date_str = self.calendar.get_date()
        file_name = date_str.replace('/', '_') + '.xlsx'
        file_path = os.path.join(self.assets_folder, file_name)

        if not os.path.exists(file_path):
            messagebox.showerror('Errore', 'File Excel non trovato per questa data.')
            return

        wb = load_workbook(file_path)
        ws = wb.active

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', size=12)
        pdf.cell(200, 10, txt=f'Report del {date_str}', ln=True, align='C')
        pdf.ln(10)
        # Inserisci una riga di separazione e il costo totale
        pdf.ln()  # Spazio vuoto
        pdf.cell(60, 10, "-------------------------", border=0)  # Linea orizzontale
        pdf.ln()  # Spazio vuoto
        pdf.cell(60, 10, "Costo Totale:", border=0)  # Etichetta 'Costo Totale'
        pdf.cell(60, 10, str(self.total_cost_var.get()), border=0)  # Costo totale (assicurati di convertire in stringa se necessario)
        pdf.ln()  # Spazio vuoto aggiuntivo (se necessario)



        headers = [cell.value for cell in ws[1]]
        for header in headers:
            pdf.cell(60, 10, header, border=1)
        pdf.ln()

        for row in ws.iter_rows(min_row=2, values_only=True):
            for item in row:
                pdf.cell(60, 10, str(item), border=1)
            pdf.ln()

        # Esporta sul Desktop (Windows o Mac)
        if os.name == 'nt':  # Windows
            desktop_path = os.path.expanduser('~\\Desktop')
        else:  # macOS o altri sistemi
            desktop_path = os.path.expanduser('~/Desktop')

        output_path = os.path.join(desktop_path, f'Report_{date_str.replace("/", "_")}.pdf')
        pdf.output(output_path)

        messagebox.showinfo('Esportazione PDF', f'PDF esportato correttamente in {output_path}')

    def update_full_item_list(self):
        self.full_item_list = {self.tree.item(row)['values'][0].lower(): row for row in self.tree.get_children()}

    def filter_users(self, event=None):
        search_term = self.search_var.get().lower()

    # Se non hai ancora salvato tutti gli elementi, lo fai ora
        if not hasattr(self, 'full_item_list'):
            self.update_full_item_list()

    # Ripulisci la tabella per evitare duplicati
        for row in self.tree.get_children():
           self.tree.detach(row)

    # Se la ricerca è vuota, riattacca tutti gli elementi
        if not search_term:
            for name, row_id in self.full_item_list.items():
                self.tree.reattach(row_id, '', 'end')
            return

    # Filtra e riattacca gli elementi che corrispondono
        for name, row_id in self.full_item_list.items():
            if search_term in name:
                self.tree.reattach(row_id, '', 'end')


    def set_date(self):
        date_str = self.calendar.get_date()
        self.current_date = datetime.strptime(date_str, '%d/%m/%Y')
        file_name = self.current_date.strftime('%d_%m_%Y') + '.xlsx'
        file_path = os.path.join(self.assets_folder, file_name)

        # Pulire la tabella prima di caricare nuovi dati
        for row in self.tree.get_children():
            self.tree.delete(row)

        if os.path.exists(file_path):
            self.load_existing_data(file_path)
            messagebox.showinfo("Caricamento", f"Dati trovati per la data selezionata.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Dati Timer"
            ws.append(['Nome', 'Timer', 'Costo'])
            wb.save(file_path)
            messagebox.showinfo("Data Confermata", f"Nuovo file Excel creato per la data: {date_str}")
    
    def load_existing_data(self, file_path):
        """Carica i dati dal file Excel nella tabella"""
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Legge i dati (salta la prima riga con i titoli)
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.tree.insert('', 'end', values=row)


    def add_user(self):
        name = self.user_name_var.get()
        if not name or not self.current_date:
            messagebox.showerror("Errore", "Devi inserire un nome e confermare una data!")
            return
    
        start_time = datetime.now()
        self.users.append({'name': name, 'start_time': start_time})
        self.running_timers[name] = start_time
        self.tree.insert('', 'end', values=(name, "00:00:00", "0.00"))
        self.user_name_var.set('')

    # Aggiorna la lista completa dopo l'inserimento
        self.update_full_item_list()



    def update_timers(self):
    # Aggiorna i timer per tutti gli utenti in esecuzione
        for user in self.users:
            name = user['name']
            start_time = self.running_timers.get(name)

            if start_time:
            # Calcola il tempo trascorso
                elapsed = datetime.now() - start_time
                hours, remainder = divmod(elapsed.seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                timer_str = f"{hours:02}:{minutes:02}:{seconds:02}"

            # Calcola il costo in base al tempo trascorso
                if elapsed <= timedelta(minutes=30):
                    cost = 5
                elif elapsed <= timedelta(minutes=60):
                    cost = 9
                elif elapsed <= timedelta(hours=1, minutes=30):
                    cost = 13
                elif elapsed <= timedelta(hours=2):
                    cost = 17
                else:
                    cost = 20

            # Aggiorna la tabella con i nuovi valori
                for row in self.tree.get_children():
                    if self.tree.item(row)['values'][0] == name:
                        self.tree.item(row, values=(name, timer_str, f"{cost:.2f}"))
                        self.tree.item(row, tags=('running',))  # Applica il tag per il colore rosso

    # Calcola il costo totale dinamico
        total_cost = 0.0
        for row in self.tree.get_children():
            cost = float(self.tree.item(row)['values'][2])
            total_cost += cost

    # Aggiorna il valore nella label
        self.total_cost_var.set(f"{total_cost:.2f}")

    # Pianifica l'aggiornamento del timer dopo 1 secondo
        self.root.after(1000, self.update_timers)


    def stop_timer(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Errore", "Seleziona un utente per fermare il timer.")
            return

        selected_name = self.tree.item(selected_item)['values'][0]

        if selected_name in self.running_timers:
        # Memorizza il tempo trascorso prima della pausa
            start_time = self.running_timers[selected_name]
            elapsed = datetime.now() - start_time

        # Aggiungi il tempo trascorso a paused_timers
            if selected_name in self.paused_timers:
                self.paused_timers[selected_name] += elapsed
            else:
                self.paused_timers[selected_name] = elapsed

            self.tree.item(selected_item, tags=())
            self.running_timers.pop(selected_name, None)

        elif selected_name in self.paused_timers:
            messagebox.showinfo("Timer fermato", "Il timer per questo utente è già fermo.")
        else:
            messagebox.showerror("Errore", "Timer non trovato per questo utente.")


    def save_to_excel(self):
        if not self.current_date:
            messagebox.showerror("Errore", "Seleziona una data prima di salvare.")
            return
        
        file_name = self.current_date.strftime('%d_%m_%Y') + '.xlsx'
        file_path = os.path.join(self.assets_folder, file_name)

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
        else:
            wb = Workbook()

        ws = wb.active
        ws.title = "Dati Timer"

        # Recupera i nomi già presenti nel file Excel
        existing_names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            existing_names.add(row[0])  # Aggiunge solo il nome alla lista
        
        # Inserisce solo i nuovi nomi non presenti
        for row in self.tree.get_children():
            data = self.tree.item(row)['values']
            if data[0] not in existing_names:
                ws.append(data)
        
        wb.save(file_path)
        messagebox.showinfo("Salvataggio", f"Dati salvati correttamente in {file_path}.")
    
    def delete_user(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showerror("Errore", "Seleziona un utente da eliminare.")
            return

        selected_name = self.tree.item(selected_item)['values'][0]
        self.tree.delete(selected_item)
        self.users = [user for user in self.users if user['name'] != selected_name]
        self.running_timers.pop(selected_name, None)
    
    # Rimozione dal file Excel
        if self.current_date:
            file_name = self.current_date.strftime('%d_%m_%Y') + '.xlsx'
            file_path = os.path.join(self.assets_folder, file_name)

            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                ws = wb.active

            # Cerca il nome e lo elimina
                for row in ws.iter_rows(min_row=2, values_only=False):
                    if row[0].value == selected_name:
                        ws.delete_rows(row[0].row, 1)
                        break

                wb.save(file_path)

        messagebox.showinfo("Eliminazione", f"Utente '{selected_name}' eliminato correttamente.")


# Avvio dell'app
root = ctk.CTk()
app = MultiUserTimerApp(root)
root.mainloop()
